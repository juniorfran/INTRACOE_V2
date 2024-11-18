from django.core.mail import EmailMessage
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from .models import Departamentos, Cargo, Empleados, Boleta_pago
from .forms import DepartamentoForm, CargoForm, EmpleadoForm, BoletaPagoForm, UploadFileForm, UploadFileFormEmpleados
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
import openpyxl
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib.auth.decorators import login_required

# Listar Departamentos
@login_required
def departamento_list(request):
    departamentos = Departamentos.objects.all()
    return render(request, 'departamentos/departamentos_list.html', {'departamentos': departamentos})

# Crear Departamento
@login_required
def departamento_create(request):
    if request.method == 'POST':
        form = DepartamentoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('departamento_list')
    else:
        form = DepartamentoForm()
    return render(request, 'departamentos/departamento_form.html', {'form': form})

# Editar Departamento
@login_required
def departamento_update(request, pk):
    departamento = get_object_or_404(Departamentos, pk=pk)
    if request.method == 'POST':
        form = DepartamentoForm(request.POST, instance=departamento)
        if form.is_valid():
            form.save()
            return redirect('departamento_list')
    else:
        form = DepartamentoForm(instance=departamento)
    return render(request, 'departamentos/departamento_form.html', {'form': form})

# Eliminar Departamento
@login_required
def departamento_delete(request, pk):
    departamento = get_object_or_404(Departamentos, pk=pk)
    if request.method == 'POST':
        departamento.delete()
        return redirect('departamento_list')
    return render(request, 'departamentos/departamento_confirm_delete.html', {'departamento': departamento})

#######################################################################################################333

# Listar Cargos
@login_required
def cargo_list(request):
    cargos = Cargo.objects.all()
    return render(request, 'cargos/cargos_list.html', {'cargos': cargos})

# Crear Cargo
@login_required
def cargo_create(request):
    if request.method == 'POST':
        form = CargoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('cargo_list')
    else:
        form = CargoForm()
    return render(request, 'cargos/cargo_form.html', {'form': form})

# Editar Cargo
@login_required
def cargo_update(request, pk):
    cargo = get_object_or_404(Cargo, pk=pk)
    if request.method == 'POST':
        form = CargoForm(request.POST, instance=cargo)
        if form.is_valid():
            form.save()
            return redirect('cargo_list')
    else:
        form = CargoForm(instance=cargo)
    return render(request, 'cargos/cargo_form.html', {'form': form})

# Eliminar Cargo
@login_required
def cargo_delete(request, pk):
    cargo = get_object_or_404(Cargo, pk=pk)
    if request.method == 'POST':
        cargo.delete()
        return redirect('cargo_list')
    return render(request, 'cargos/cargo_confirm_delete.html', {'cargo': cargo})



#######################################################################################################333

# Listar Boletas de pago con búsqueda por código de empleado
@login_required
def boleta_pago_list(request):
    # Obtener parámetros de búsqueda
    codigo_empleado = request.GET.get('codigo_empleado', '')
    fecha_pago = request.GET.get('fecha_pago', '')

    # Filtrar boletas según el código de empleado y fecha de pago
    boletas = Boleta_pago.objects.all()

    if codigo_empleado:
        boletas = boletas.filter(empleado__codigo_empleado__icontains=codigo_empleado)

    if fecha_pago:
        boletas = boletas.filter(fecha_pago=fecha_pago)

    # Paginación: 25 registros por página
    paginator = Paginator(boletas, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'boletas/boleta_pago_list.html', {
        'page_obj': page_obj,
        'codigo_empleado': codigo_empleado,
        'fecha_pago': fecha_pago
    })


# Crear Boleta de pago
@login_required
def boleta_pago_create(request):
    if request.method == 'POST':
        form = BoletaPagoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('boleta_pago_list')
    else:
        form = BoletaPagoForm()
    return render(request, 'boletas/boleta_pago_form.html', {'form': form})

# Editar Boleta de pago
@login_required
def boleta_pago_update(request, pk):
    boleta = get_object_or_404(Boleta_pago, pk=pk)
    if request.method == 'POST':
        form = BoletaPagoForm(request.POST, instance=boleta)
        if form.is_valid():
            form.save()
            return redirect('boleta_pago_list')
    else:
        form = BoletaPagoForm(instance=boleta)
    return render(request, 'boletas/boleta_pago_form.html', {'form': form})

# Eliminar Boleta de pago
@login_required
def boleta_pago_delete(request, pk):
    boleta = get_object_or_404(Boleta_pago, pk=pk)
    if request.method == 'POST':
        boleta.delete()
        return redirect('boleta_pago_list')
    return render(request, 'boletas/boleta_pago_confirm_delete.html', {'boleta': boleta})

#######################################################################################################333

# Listar Empleados
@login_required
def empleado_list(request):
    empleados = Empleados.objects.all()
    return render(request, 'empleados/empleados_list.html', {'empleados': empleados})

# Crear Empleado
@login_required
def empleado_create(request):
    if request.method == 'POST':
        form = EmpleadoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('empleado_list')
    else:
        form = EmpleadoForm()
    return render(request, 'empleados/empleado_form.html', {'form': form})

# Editar Empleado
@login_required
def empleado_update(request, pk):
    empleado = get_object_or_404(Empleados, pk=pk)
    if request.method == 'POST':
        form = EmpleadoForm(request.POST, instance=empleado)
        if form.is_valid():
            form.save()
            return redirect('empleado_list')
    else:
        form = EmpleadoForm(instance=empleado)
    return render(request, 'empleados/empleado_form.html', {'form': form})

# Eliminar Empleado
@login_required
def empleado_delete(request, pk):
    empleado = get_object_or_404(Empleados, pk=pk)
    if request.method == 'POST':
        empleado.delete()
        return redirect('empleado_list')
    return render(request, 'empleados/empleado_confirm_delete.html', {'empleado': empleado})

@login_required
def cargar_empleados_desde_xlsx(request):
    if request.method == 'POST':
        form = UploadFileFormEmpleados(request.POST, request.FILES)
        
        if form.is_valid():
            archivo_xlsx = request.FILES['file']
            wb = openpyxl.load_workbook(archivo_xlsx)
            hoja = wb.active  # Usamos la primera hoja del archivo
            
            # Iterar sobre las filas (omitiendo la primera si contiene encabezados)
            for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
                nombre, apellido, dui, codigo_empleado, edad, salario, cargo, email, num_telefono = fila
                

                # Verificar si el cargo existe o se crea
                #cargo, created = Cargo.objects.get_or_create(nombre=cargo_nombre)
                
                # Crear o actualizar el empleado
                empleado, creado = Empleados.objects.update_or_create(
                    codigo_empleado=codigo_empleado,  # Campo único
                    defaults={
                        'nombre': nombre,
                        'apellido': apellido,
                        'dui': dui,
                        'edad': edad,
                        'salario': salario,
                        'cargo': cargo,
                        'email': email,
                        'num_telefono': num_telefono,
                    }
                )
                
                action = "Creado" if creado else "Actualizado"
                print(f"{action} empleado en la fila {i}: {empleado.nombre} {empleado.apellido}")

            messages.success(request, "Los empleados han sido cargados y guardados correctamente.")
            return redirect('empleado_list')  # Redirige a la lista de empleados
    else:
        form = UploadFileFormEmpleados()
    
    return render(request, 'cargar_empleados.html', {'form': form})

###############################################################################################
@login_required
def cargar_boletas(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['file']
            try:
                workbook = openpyxl.load_workbook(archivo)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorar encabezado
                    empleado_codigo = row[0]
                    empleado = Empleados.objects.get(codigo_empleado=empleado_codigo)

                    Boleta_pago.objects.create(
                        fecha_pago=row[1],
                        fecha_inicio=row[2],
                        fecha_fin=row[3],
                        dias_laborados=row[4],
                        empleado=empleado,
                        descuento_afp=row[5],
                        descuento_isss=row[6],
                        descuento_renta=row[7],
                        otro_descuento1=row[8],
                        otro_descuento2=row[9],
                        total_descuentos=row[10],
                        hr_extra_di=row[11],
                        hr_extra_noc=row[12],
                        hr_extra_fer=row[13],
                        hr_extra_fer_noc=row[14],
                        total_pago=row[15],
                        liquido_recibir=row[16],

                    )

                messages.success(request, "Las boletas de pago se han cargado exitosamente.")
                return redirect('boleta_pago_list')

            except Exception as e:
                messages.error(request, f"Hubo un error al procesar el archivo: {e}")
        else:
            messages.error(request, "Por favor seleccione un archivo válido.")
    else:
        form = UploadFileForm()
    return render(request, 'cargar_boletas.html', {'form': form})


###############################################################################################

from xhtml2pdf import pisa
from io import BytesIO
from django.utils import timezone

@login_required
def enviar_boletas_masivo(request):
    # Obtener la fecha de pago más reciente
    boletas_recientes = Boleta_pago.objects.order_by('-fecha_pago').first()
    if boletas_recientes:
        fecha_pago_mas_reciente = boletas_recientes.fecha_pago

        # Filtrar boletas por la fecha de pago más reciente
        boletas = Boleta_pago.objects.filter(fecha_pago=fecha_pago_mas_reciente)
        
        print(f"Cantidad de boletas a enviar: {boletas.count()}")

        for boleta in boletas:
            empleado = boleta.empleado
            print(f"Enviando boleta a: {empleado.nombre} {empleado.apellido} - Email: {empleado.email}")
            
            # Renderizar el HTML de la boleta
            salario_quince = empleado.salario / 2
            html_content = render_to_string('boleta_pago_pdf.html', {'boleta': boleta, 'empleado': empleado, 'salario_quince': salario_quince})
            
            # Convertir el HTML a PDF
            pdf_file = BytesIO()
            pisa_status = pisa.CreatePDF(BytesIO(html_content.encode('UTF-8')), dest=pdf_file)
            
            if pisa_status.err:
                print(f"Error al crear el PDF para {empleado.nombre} {empleado.apellido}")
                continue

            pdf_file.seek(0)  # Volver al inicio del archivo PDF

            # Renderizar el HTML del mensaje del correo
            email_html_content = render_to_string('boleta_pago_email.html', {'boleta': boleta, 'empleado': empleado, 'salario_quince': salario_quince})
            
            # Crear el correo electrónico con formato HTML
            email = EmailMessage(
                subject="Boleta de Pago",
                body=email_html_content,
                from_email=settings.EMAIL_HOST_USER,
                to=[empleado.email],
            )
            email.content_subtype = "html"  # Indicar que el contenido es HTML
            email.attach(f"Boleta_Pago_{empleado.nombre}_{empleado.apellido}.pdf", pdf_file.read(), 'application/pdf')

            try:
                email.send(fail_silently=False)
                print(f"Boleta enviada a {empleado.email}")
            
            except Exception as e:
                print(f"Error al enviar boleta a {empleado.email}: {e}")

        messages.success(request, "Las boletas de pago han sido enviadas a todos los empleados con la fecha de pago más reciente.")
    else:
        messages.warning(request, "No se encontraron boletas de pago para enviar.")

    return redirect('boleta_pago_list')

###############################################################################################

@login_required
def enviar_boleta_individual(request, empleado_id):
    empleado = get_object_or_404(Empleados, id=empleado_id)
    boleta = Boleta_pago.objects.filter(empleado=empleado).order_by('-fecha_pago').first()  # Última boleta de pago

    if boleta:
        print(f"Enviando boleta a: {empleado.nombre} {empleado.apellido} - Email: {empleado.email}")
        
        # Renderizar el HTML de la boleta
        salario_quince = empleado.salario / 2
        html_content = render_to_string('boleta_pago_pdf.html', {'boleta': boleta, 'empleado': empleado, 'salario_quince': salario_quince})
        
        # Convertir el HTML a PDF
        pdf_file = BytesIO()
        pisa_status = pisa.CreatePDF(BytesIO(html_content.encode('UTF-8')), dest=pdf_file)
        
        if pisa_status.err:
            print(f"Error al crear el PDF para {empleado.nombre} {empleado.apellido}")
            messages.error(request, f"Error al crear el PDF para {empleado.nombre} {empleado.apellido}")
            return redirect('boleta_pago_list')

        pdf_file.seek(0)  # Volver al inicio del archivo PDF

        # Renderizar el HTML del mensaje del correo
        email_html_content = render_to_string('boleta_pago_email.html', {'boleta': boleta, 'empleado': empleado, 'salario_quince': salario_quince})
        
        # Crear el correo electrónico con formato HTML
        email = EmailMessage(
            subject="Boleta de Pago",
            body=email_html_content,
            from_email=settings.EMAIL_HOST_USER,
            to=[empleado.email],
        )
        email.content_subtype = "html"  # Indicar que el contenido es HTML
        email.attach(f"Boleta_Pago_{empleado.nombre}_{empleado.apellido}.pdf", pdf_file.read(), 'application/pdf')

        try:
            email.send(fail_silently=False)
            print(f"Boleta enviada a {empleado.email}")
            messages.success(request, f"La boleta de pago fue enviada exitosamente a {empleado.nombre} {empleado.apellido}")
        
        except Exception as e:
            print(f"Error al enviar boleta a {empleado.email}: {e}")
            messages.error(request, f"Error al enviar la boleta a {empleado.email}: {e}")
    else:
        messages.warning(request, f"No se encontró una boleta de pago para {empleado.nombre} {empleado.apellido}.")

    return redirect('boleta_pago_list')


########################################################################################################################################3

@login_required
def acciones_boleta_pago_list(request):
    # Obtener parámetros de búsqueda
    codigo_empleado = request.GET.get('codigo_empleado', '')
    fecha_pago = request.GET.get('fecha_pago', '')

    # Filtrar boletas según el código de empleado y fecha de pago
    boletas = Boleta_pago.objects.all()

    if codigo_empleado:
        boletas = boletas.filter(empleado__codigo_empleado__icontains=codigo_empleado)

    if fecha_pago:
        boletas = boletas.filter(fecha_pago=fecha_pago)

    # Paginación: 25 registros por página
    paginator = Paginator(boletas, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'boletas/acciones_boleta_pago_list.html', {
        'page_obj': page_obj,
        'codigo_empleado': codigo_empleado,
        'fecha_pago': fecha_pago,
    })

@login_required
def acciones_enviar_boletas_masivo(request):
    # Obtener IDs de las boletas seleccionadas
    boletas_ids = request.POST.getlist('boleta_ids')
    
    if not boletas_ids:
        messages.warning(request, "No se seleccionaron boletas para enviar.")
        return redirect('acciones_boleta_pago_list')

    boletas = Boleta_pago.objects.filter(id__in=boletas_ids)
    print(f"Cantidad de boletas a enviar: {boletas.count()}")

    for boleta in boletas:
        empleado = boleta.empleado
        messages.warning(request, f"Enviando boleta a: {empleado.nombre} {empleado.apellido} - Email: {empleado.email}")
        print(f"Enviando boleta a: {empleado.nombre} {empleado.apellido} - Email: {empleado.email}")
        
        # Renderizar el HTML de la boleta
        salario_quince = empleado.salario / 2
        html_content = render_to_string('boleta_pago_pdf.html', {'boleta': boleta, 'empleado': empleado, 'salario_quince': salario_quince})
        
        # Convertir el HTML a PDF
        pdf_file = BytesIO()
        pisa_status = pisa.CreatePDF(BytesIO(html_content.encode('UTF-8')), dest=pdf_file)
        
        if pisa_status.err:
            print(f"Error al crear el PDF para {empleado.nombre} {empleado.apellido}")
            continue

        pdf_file.seek(0)  # Volver al inicio del archivo PDF

        # Renderizar el HTML del mensaje del correo
        email_html_content = render_to_string('boleta_pago_email.html', {'boleta': boleta, 'empleado': empleado, 'salario_quince': salario_quince})
        
        # Crear el correo electrónico con formato HTML
        email = EmailMessage(
            subject="Boleta de Pago",
            body=email_html_content,
            from_email=settings.EMAIL_HOST_USER,
            to=[empleado.email],
        )
        email.content_subtype = "html"  # Indicar que el contenido es HTML
        email.attach(f"Boleta_Pago_{empleado.nombre}_{empleado.apellido}.pdf", pdf_file.read(), 'application/pdf')

        try:
            email.send(fail_silently=False)
            print(f"Boleta enviada a {empleado.email}")
        
        except Exception as e:
            messages.error(request, f"Error al enviar las boletas a {empleado.email}:{e}")
            print(f"Error al enviar boleta a {empleado.email}: {e}")

    messages.success(request, "Las boletas seleccionadas han sido enviadas exitosamente.")
    return redirect('acciones_boleta_pago_list')





